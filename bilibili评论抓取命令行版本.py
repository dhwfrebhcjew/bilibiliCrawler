#!/usr/bin/env python3
"""
B站评论爬虫 - 增强稳定版（指数退避 + 智能延迟）
"""
import requests
import json
import re
import time
import os
from datetime import datetime
import random

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

class Colors:
    """终端颜色"""
    PINK = '\033[38;5;205m'
    BLUE = '\033[38;5;39m'
    GREEN = '\033[38;5;46m'
    YELLOW = '\033[38;5;226m'
    RED = '\033[38;5;196m'
    CYAN = '\033[38;5;51m'
    PURPLE = '\033[38;5;129m'
    WHITE = '\033[38;5;255m'
    GRAY = '\033[38;5;242m'
    BOLD = '\033[1m'
    RESET = '\033[0m'
    CLEAR = '\033[2J\033[H'

def clear_screen():
    print(Colors.CLEAR, end='')

def print_banner():
    banner = f"""
{Colors.PINK}{Colors.BOLD}
╔══════════════════════════════════════════════════════════╗
║         B站评论爬虫 Pro v3.2 - 智能防封版                ║
║      指数退避 | 随机延迟2-5秒 | 自动重试                 ║
╚══════════════════════════════════════════════════════════╝
{Colors.RESET}"""
    print(banner)

def print_status(msg, status='info'):
    icons = {
        'info': f'{Colors.BLUE}ℹ{Colors.RESET}',
        'success': f'{Colors.GREEN}✅{Colors.RESET}',
        'error': f'{Colors.RED}❌{Colors.RESET}',
        'warning': f'{Colors.YELLOW}⚠️{Colors.RESET}',
        'loading': f'{Colors.PINK}⏳{Colors.RESET}',
    }
    print(f"\n{icons.get(status, '')} {msg}")

def print_progress(current, total, prefix=''):
    bar_length = 30
    filled = int(bar_length * current / total) if total > 0 else 0
    bar = f"{Colors.PINK}{'█' * filled}{Colors.GRAY}{'░' * (bar_length - filled)}{Colors.RESET}"
    percent = f"{current/total*100:.1f}%" if total > 0 else "0%"
    print(f"\r{prefix} {bar} {Colors.BOLD}{percent}{Colors.RESET} ({current}/{total})", end='')

class BilibiliComment:
    def __init__(self, cookie_file='bilibili_cookie.txt'):
        self.session = requests.Session()
        self.cookie = self.load_cookie(cookie_file)
        
        # User-Agent池，避免被识别
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0',
        ]
        
        self.update_headers()
        
        if self.cookie:
            self.session.headers['Cookie'] = self.cookie
        
        # 请求计数，用于动态调整延迟
        self.request_count = 0
        self.last_request_time = 0
    
    def update_headers(self):
        """随机更换User-Agent"""
        self.session.headers.update({
            'User-Agent': random.choice(self.user_agents),
            'Referer': 'https://www.bilibili.com',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache',
        })
    
    def exponential_backoff(self, attempt, base_delay=5, max_delay=300):
        """指数退避算法"""
        delay = min(base_delay * (2 ** attempt), max_delay)
        # 添加随机抖动，避免同时重试
        jitter = random.uniform(0, delay * 0.3)
        total_delay = delay + jitter
        return total_delay
    
    def smart_delay(self):
        """智能延迟：2-5秒随机延迟，并根据请求次数动态调整"""
        # 基础延迟2-5秒
        base_delay = random.uniform(2, 5)
        
        # 每50次请求后，增加额外延迟（模拟人类行为）
        if self.request_count > 0 and self.request_count % 50 == 0:
            extra_delay = random.uniform(3, 8)
            total_delay = base_delay + extra_delay
            print(f"\n{Colors.GRAY}📊 已完成{self.request_count}次请求，休息{extra_delay:.1f}秒{Colors.RESET}")
        else:
            total_delay = base_delay
        
        # 确保与上次请求至少间隔1秒
        if self.last_request_time > 0:
            elapsed = time.time() - self.last_request_time
            if elapsed < 1:
                time.sleep(1 - elapsed)
        
        time.sleep(total_delay)
        self.last_request_time = time.time()
        self.request_count += 1
        
        return total_delay
    
    def load_cookie(self, filepath):
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                cookie = f.read().strip()
            return cookie
        return None
    
    def check_login(self):
        if not self.cookie:
            return None
        
        try:
            resp = self.session.get('https://api.bilibili.com/x/web-interface/nav', timeout=10)
            data = resp.json()
            if data['code'] == 0 and data['data']['isLogin']:
                return {
                    'name': data['data']['uname'],
                    'level': data['data']['level_info']['current_level'],
                    'coins': data['data']['money'],
                }
        except:
            pass
        return None
    
    def get_video_info(self, url):
        bv = re.search(r'BV[a-zA-Z0-9]+', url)
        bvid = bv.group() if bv else url
        
        for retry in range(3):
            try:
                self.update_headers()  # 更换UA
                resp = self.session.get(f'https://api.bilibili.com/x/web-interface/view?bvid={bvid}', timeout=10)
                
                if resp.status_code == 412:
                    delay = self.exponential_backoff(retry)
                    print_status(f"请求被限制，等待 {delay:.1f} 秒后重试...", 'warning')
                    time.sleep(delay)
                    continue
                
                data = resp.json()
                
                if data['code'] == 0:
                    return {
                        'aid': data['data']['aid'],
                        'title': data['data']['title'],
                        'total': data['data']['stat']['reply'],
                        'like': data['data']['stat']['like'],
                        'view': data['data']['stat']['view'],
                        'danmaku': data['data']['stat']['danmaku'],
                    }
                else:
                    print_status(f"API错误: {data.get('message', '未知错误')}", 'error')
                    
            except Exception as e:
                print_status(f"获取视频信息失败 (重试{retry+1}/3): {e}", 'warning')
                time.sleep(self.exponential_backoff(retry))
        
        return None
    
    def fetch_comments_safe(self, url, max_retries=5):
        """安全的API请求，带指数退避重试机制"""
        for attempt in range(max_retries):
            try:
                # 随机更换User-Agent
                self.update_headers()
                
                resp = self.session.get(url, timeout=15)
                
                # 检查状态码
                if resp.status_code == 403 or resp.status_code == 412:
                    delay = self.exponential_backoff(attempt)
                    print(f"\n{Colors.YELLOW}⚠️ 被限制访问 (状态码{resp.status_code})，等待 {delay:.1f} 秒后重试{Colors.RESET}")
                    time.sleep(delay)
                    continue
                
                if resp.status_code != 200:
                    if attempt < max_retries - 1:
                        delay = self.exponential_backoff(attempt)
                        print(f"\n{Colors.YELLOW}⚠️ HTTP {resp.status_code}，等待 {delay:.1f} 秒后重试{Colors.RESET}")
                        time.sleep(delay)
                        continue
                    return None
                
                # 检查是否为空响应
                if not resp.text or not resp.text.strip():
                    if attempt < max_retries - 1:
                        delay = self.exponential_backoff(attempt)
                        print(f"\n{Colors.YELLOW}⚠️ 空响应，等待 {delay:.1f} 秒后重试{Colors.RESET}")
                        time.sleep(delay)
                        continue
                    return None
                
                # 尝试解析JSON
                data = resp.json()
                
                # 检查是否是反爬虫页面
                if isinstance(data, dict) and data.get('code') == -403:
                    print(f"\n{Colors.RED}❌ 被反爬虫拦截{Colors.RESET}")
                    if attempt < max_retries - 1:
                        delay = self.exponential_backoff(attempt, base_delay=10)
                        print(f"{Colors.YELLOW}等待 {delay:.1f} 秒后重试...{Colors.RESET}")
                        time.sleep(delay)
                        continue
                    return None
                
                return data
                
            except json.JSONDecodeError as e:
                print(f"\n{Colors.YELLOW}⚠️ JSON解析失败 (尝试 {attempt+1}/{max_retries}){Colors.RESET}")
                if 'resp' in locals() and resp.text:
                    print(f"{Colors.GRAY}响应内容前100字符: {resp.text[:100]}{Colors.RESET}")
                if attempt < max_retries - 1:
                    delay = self.exponential_backoff(attempt)
                    time.sleep(delay)
                else:
                    return None
                    
            except requests.Timeout:
                print(f"\n{Colors.YELLOW}⚠️ 请求超时 (尝试 {attempt+1}/{max_retries}){Colors.RESET}")
                if attempt < max_retries - 1:
                    delay = self.exponential_backoff(attempt)
                    time.sleep(delay)
                    continue
                return None
                
            except requests.RequestException as e:
                print(f"\n{Colors.YELLOW}⚠️ 网络错误 (尝试 {attempt+1}/{max_retries}): {e}{Colors.RESET}")
                if attempt < max_retries - 1:
                    delay = self.exponential_backoff(attempt)
                    time.sleep(delay)
                else:
                    return None
        
        return None
    
    def fetch_comments(self, aid, max_count=500, sort=1, keyword=None, progress_callback=None):
        """爬取指定数量评论"""
        comments = []
        filtered = []
        page = 1
        failed_pages = 0
        max_failed = 5
        consecutive_empty = 0
        
        print(f"{Colors.CYAN}开始爬取评论...{Colors.RESET}")
        print(f"{Colors.GRAY}策略: 每次请求后延迟2-5秒，被限制时指数退避{Colors.RESET}\n")
        
        while len(comments) < max_count:
            params = {
                'type': 1,
                'oid': aid,
                'pn': page,
                'ps': 20,
                'sort': sort,
            }
            
            url = 'https://api.bilibili.com/x/v2/reply'
            for key, value in params.items():
                url += f'&{key}={value}' if '?' in url else f'?{key}={value}'
            
            result = self.fetch_comments_safe(url)
            
            if result is None:
                failed_pages += 1
                if failed_pages >= max_failed:
                    print_status("连续失败次数过多，停止爬取", 'error')
                    break
                page += 1
                continue
            
            failed_pages = 0
            
            if result.get('code') != 0:
                error_msg = result.get('message', '未知错误')
                print(f"\n{Colors.YELLOW}⚠️ API返回错误: {error_msg}{Colors.RESET}")
                break
            
            replies = result.get('data', {}).get('replies', [])
            if not replies:
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    break
                page += 1
                continue
            else:
                consecutive_empty = 0
            
            # 处理评论
            for r in replies:
                if len(comments) >= max_count:
                    break
                    
                content = r.get('content', {}).get('message', '')
                if not content or not content.strip():
                    continue
                
                comment_data = {
                    '序号': len(comments) + 1,
                    '用户': r.get('member', {}).get('uname', '未知'),
                    '用户ID': r.get('member', {}).get('mid', 0),
                    '等级': r.get('member', {}).get('level_info', {}).get('current_level', 0),
                    '内容': content,
                    '点赞': r.get('like', 0),
                    '回复数': r.get('rcount', 0),
                    '时间': datetime.fromtimestamp(r.get('ctime', 0)).strftime('%Y-%m-%d %H:%M:%S'),
                }
                comments.append(comment_data)
                
                if keyword and keyword.lower() in content.lower():
                    filtered.append(comment_data)
            
            if progress_callback:
                progress_callback(len(comments), max_count, page)
            
            page += 1
            
            # 智能延迟2-5秒
            delay = self.smart_delay()
            print(f"\r{Colors.GRAY}💤 等待 {delay:.1f} 秒后继续...{Colors.RESET}", end='')
        
        print()  # 换行
        return comments, filtered
    
    def fetch_all_comments(self, aid, sort=1, keyword=None, progress_callback=None, stop_callback=None):
        """爬取所有评论（无上限，带指数退避）"""
        comments = []
        filtered = []
        page = 1
        failed_pages = 0
        max_failed = 8
        consecutive_empty = 0
        total_comments = None
        
        print(f"{Colors.CYAN}开始爬取所有评论...{Colors.RESET}")
        print(f"{Colors.GRAY}策略: 每页延迟2-5秒，被限制时指数退避{Colors.RESET}\n")
        
        while True:
            params = {
                'type': 1,
                'oid': aid,
                'pn': page,
                'ps': 20,
                'sort': sort,
            }
            
            url = 'https://api.bilibili.com/x/v2/reply'
            for key, value in params.items():
                url += f'&{key}={value}' if '?' in url else f'?{key}={value}'
            
            result = self.fetch_comments_safe(url)
            
            if result is None:
                failed_pages += 1
                print(f"\n{Colors.YELLOW}⚠️ 第{page}页请求失败 ({failed_pages}/{max_failed}){Colors.RESET}")
                if failed_pages >= max_failed:
                    print_status("连续失败次数过多，停止爬取", 'error')
                    break
                page += 1
                continue
            
            failed_pages = 0
            
            if result.get('code') != 0:
                error_msg = result.get('message', '未知错误')
                print(f"\n{Colors.YELLOW}⚠️ API返回错误: {error_msg}{Colors.RESET}")
                if result.get('code') in [-404, -400]:
                    print(f"{Colors.GRAY}已到达最后一页{Colors.RESET}")
                    break
                page += 1
                continue
            
            # 获取总评论数
            if total_comments is None:
                total_comments = result.get('data', {}).get('page', {}).get('count', 0)
                if total_comments > 0:
                    print(f"{Colors.GREEN}📊 总评论数: {total_comments:,}{Colors.RESET}")
                else:
                    print(f"{Colors.YELLOW}⚠️ 无法获取总评论数，继续爬取...{Colors.RESET}")
            
            replies = result.get('data', {}).get('replies', [])
            
            if not replies:
                consecutive_empty += 1
                if consecutive_empty >= 5:
                    print(f"\n{Colors.GRAY}已到达最后一页{Colors.RESET}")
                    break
                print(f"\n{Colors.GRAY}第{page}页无评论数据{Colors.RESET}")
                page += 1
                continue
            else:
                consecutive_empty = 0
            
            # 处理评论
            new_comments = 0
            for r in replies:
                content = r.get('content', {}).get('message', '')
                if not content or not content.strip():
                    continue
                
                comment_data = {
                    '序号': len(comments) + 1,
                    '用户': r.get('member', {}).get('uname', '未知'),
                    '用户ID': r.get('member', {}).get('mid', 0),
                    '等级': r.get('member', {}).get('level_info', {}).get('current_level', 0),
                    '内容': content,
                    '点赞': r.get('like', 0),
                    '回复数': r.get('rcount', 0),
                    '时间': datetime.fromtimestamp(r.get('ctime', 0)).strftime('%Y-%m-%d %H:%M:%S'),
                }
                comments.append(comment_data)
                new_comments += 1
                
                if keyword and keyword.lower() in content.lower():
                    filtered.append(comment_data)
            
            print(f"\n{Colors.GREEN}✓ 第{page}页 +{new_comments}条 (总计{len(comments)}条){Colors.RESET}")
            
            if progress_callback:
                if total_comments:
                    progress_callback(len(comments), total_comments, page)
                else:
                    progress_callback(len(comments), len(comments) + 20, page)
            
            # 检查是否还有更多页
            total_pages = (result.get('data', {}).get('page', {}).get('count', 0) + 19) // 20
            if page >= total_pages and total_pages > 0:
                print(f"\n{Colors.GREEN}✅ 已获取全部 {total_pages} 页{Colors.RESET}")
                break
            
            if stop_callback and stop_callback():
                print(f"\n{Colors.YELLOW}⏸️ 用户停止了爬取{Colors.RESET}")
                break
            
            page += 1
            
            # 智能延迟2-5秒
            delay = self.smart_delay()
            print(f"\r{Colors.GRAY}💤 等待 {delay:.1f} 秒后继续...{Colors.RESET}", end='')
        
        print()  # 换行
        return comments, filtered
    
    def save_excel(self, comments, title, keyword=None):
        if not HAS_EXCEL:
            print_status("请安装openpyxl: pip install openpyxl", 'warning')
            return None
        
        if not comments:
            print_status("没有数据可导出", 'warning')
            return None
        
        safe_title = re.sub(r'[\\/:*?"<>|]', '_', title)[:40]
        prefix = f"筛选_{keyword}_" if keyword else ""
        filename = f"B站评论_{prefix}{safe_title}_{len(comments)}条.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "评论数据"
        
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='FB7299', end_color='FB7299', fill_type='solid')
        
        headers = ['序号', '用户', '用户ID', '等级', '内容', '点赞数', '回复数', '时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row, c in enumerate(comments, 2):
            ws.cell(row=row, column=1, value=c['序号'])
            ws.cell(row=row, column=2, value=c['用户'])
            ws.cell(row=row, column=3, value=c['用户ID'])
            ws.cell(row=row, column=4, value=c['等级'])
            ws.cell(row=row, column=5, value=c['内容'])
            ws.cell(row=row, column=6, value=c['点赞'])
            ws.cell(row=row, column=7, value=c['回复数'])
            ws.cell(row=row, column=8, value=c['时间'])
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:H{len(comments)+1}'
        
        wb.save(filename)
        return filename
    
    def save_txt(self, comments, title, keyword=None):
        if not comments:
            return None
            
        safe_title = re.sub(r'[\\/:*?"<>|]', '_', title)[:40]
        prefix = f"筛选_{keyword}_" if keyword else ""
        filename = f"B站评论_{prefix}{safe_title}_{len(comments)}条.txt"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"视频: {title}\n")
            f.write(f"评论: {len(comments)}条\n")
            if keyword:
                f.write(f"关键词: {keyword}\n")
            f.write(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*60 + "\n\n")
            
            for c in comments:
                f.write(f"{'─'*60}\n")
                f.write(f"【{c['序号']}】{c['用户']} | Lv{c['等级']} | {c['时间']}\n")
                f.write(f"👍{c['点赞']}  💬{c['回复数']}\n")
                f.write(f"\n{c['内容']}\n\n")
        
        return filename

class App:
    def __init__(self):
        self.crawler = BilibiliComment()
        self.user_info = None
        self.video_info = None
        self.comments = []
        self.filtered = []
        self.stop_flag = False
    
    def stop_callback(self):
        return self.stop_flag
    
    def print_main_menu(self):
        """打印主菜单"""
        print(f"\n{Colors.CYAN}{Colors.BOLD}━ 主菜单 {Colors.RESET}")
        print(f"  {Colors.YELLOW}[1]{Colors.RESET} 输入视频链接")
        print(f"  {Colors.YELLOW}[2]{Colors.RESET} 查看登录状态")
        print(f"  {Colors.YELLOW}[3]{Colors.RESET} 查看已保存文件")
        print(f"  {Colors.YELLOW}[q]{Colors.RESET} 退出程序")
    
    def print_operation_menu(self):
        """打印操作菜单"""
        print(f"\n{Colors.CYAN}{Colors.BOLD}━ 操作 {Colors.RESET}")
        print(f"  {Colors.YELLOW}[1]{Colors.RESET} 开始爬取评论")
        print(f"  {Colors.YELLOW}[2]{Colors.RESET} 设置关键词筛选")
        print(f"  {Colors.YELLOW}[3]{Colors.RESET} 更换视频")
        print(f"  {Colors.YELLOW}[b]{Colors.RESET} 返回主菜单")
    
    def run(self):
        clear_screen()
        print_banner()
        
        # 检查登录状态
        self.user_info = self.crawler.check_login()
        if self.user_info:
            print_status(f"已登录: {self.user_info['name']} (Lv{self.user_info['level']})", 'success')
        else:
            print_status("未登录，建议先获取Cookie避免被限制", 'warning')
            print(f"\n{Colors.GRAY}运行 python cookie.py 获取Cookie{Colors.RESET}")
        
        while True:
            self.print_main_menu()
            choice = input(f"\n{Colors.BOLD}请选择: {Colors.RESET}").strip()
            
            if choice == '1':
                self.video_menu()
            elif choice == '2':
                self.show_login_status()
            elif choice == '3':
                self.show_files()
            elif choice == 'q':
                print(f"\n{Colors.PINK}👋 再见！{Colors.RESET}\n")
                break
            else:
                print_status("无效选项", 'warning')
                time.sleep(0.5)
    
    def video_menu(self):
        clear_screen()
        print_banner()
        
        url = input(f"\n{Colors.CYAN}🔗 输入B站视频链接或BV号: {Colors.RESET}").strip()
        if not url:
            return
        
        print_status("正在获取视频信息...", 'loading')
        self.video_info = self.crawler.get_video_info(url)
        
        if not self.video_info:
            print_status("获取失败，请检查链接或网络", 'error')
            input(f"\n{Colors.GRAY}按回车返回...{Colors.RESET}")
            return
        
        vi = self.video_info
        print(f"\n{Colors.BOLD}📺 {vi['title']}{Colors.RESET}")
        print(f"{Colors.GRAY}├ 播放: {vi['view']:,}  弹幕: {vi['danmaku']:,}{Colors.RESET}")
        print(f"{Colors.GRAY}├ 点赞: {vi['like']:,}  评论: {vi['total']:,}{Colors.RESET}")
        
        while True:
            self.print_operation_menu()
            choice = input(f"\n{Colors.BOLD}请选择: {Colors.RESET}").strip()
            
            if choice == '1':
                self.fetch_menu()
            elif choice == '2':
                self.keyword_menu()
            elif choice == '3':
                return
            elif choice == 'b':
                return
    
    def fetch_menu(self):
        clear_screen()
        print_banner()
        
        if self.video_info:
            print(f"\n📺 {self.video_info['title'][:50]}")
            print(f"💬 总评论: {self.video_info['total']:,}")
        
        # 排序
        print(f"\n{Colors.CYAN}{Colors.BOLD}━ 排序方式 {Colors.RESET}")
        print(f"  {Colors.YELLOW}[1]{Colors.RESET} 按热度")
        print(f"  {Colors.YELLOW}[2]{Colors.RESET} 按时间")
        sort_choice = input(f"\n{Colors.BOLD}选择: {Colors.RESET}").strip()
        sort = 2 if sort_choice == '2' else 1
        
        # 爬取方式
        print(f"\n{Colors.CYAN}{Colors.BOLD}━ 爬取方式 {Colors.RESET}")
        print(f"  {Colors.YELLOW}[1]{Colors.RESET} 爬取全部评论 ({self.video_info['total']:,}条)")
        print(f"  {Colors.YELLOW}[2]{Colors.RESET} 爬取指定数量")
        mode_choice = input(f"\n{Colors.BOLD}选择: {Colors.RESET}").strip()
        
        max_count = None
        if mode_choice == '2':
            num = input(f"{Colors.CYAN}爬取数量: {Colors.RESET}").strip()
            max_count = int(num) if num else 200
            print(f"{Colors.GRAY}将爬取 {max_count} 条评论{Colors.RESET}")
        else:
            print(f"{Colors.GREEN}将爬取全部 {self.video_info['total']:,} 条评论{Colors.RESET}")
        
        # 关键词
        keyword = None
        if hasattr(self, 'keyword'):
            keyword = self.keyword
            print(f"\n🔑 关键词筛选: {Colors.YELLOW}{keyword}{Colors.RESET}")
        
        print(f"\n{Colors.PINK}🔍 开始爬取...{Colors.RESET}")
        print(f"{Colors.GRAY}提示: 按 Ctrl+C 可中途停止{Colors.RESET}\n")
        
        self.stop_flag = False
        
        def progress_callback(current, total, page):
            print_progress(current, total, f"第{page}页")
        
        try:
            if max_count:
                self.comments, self.filtered = self.crawler.fetch_comments(
                    self.video_info['aid'],
                    max_count,
                    sort,
                    keyword,
                    progress_callback
                )
            else:
                self.comments, self.filtered = self.crawler.fetch_all_comments(
                    self.video_info['aid'],
                    sort,
                    keyword,
                    progress_callback,
                    self.stop_callback
                )
        except KeyboardInterrupt:
            print(f"\n\n{Colors.YELLOW}⚠️ 用户中断爬取{Colors.RESET}")
            self.stop_flag = True
        
        print()  # 换行
        
        if self.comments:
            show_data = self.filtered if keyword else self.comments
            count = len(show_data)
            
            print_status(f"爬取完成！共 {len(self.comments):,} 条", 'success')
            if keyword:
                print_status(f"匹配关键词: {count:,} 条", 'info')
            
            # 显示统计
            if show_data:
                likes = [c['点赞'] for c in show_data]
                print(f"\n{Colors.BOLD}📊 统计:{Colors.RESET}")
                print(f"  最多赞: {Colors.PINK}{max(likes):,}{Colors.RESET}")
                print(f"  平均赞: {Colors.YELLOW}{sum(likes)//len(likes):,}{Colors.RESET}")
            
            # 导出
            print(f"\n{Colors.CYAN}{Colors.BOLD}━ 导出格式 {Colors.RESET}")
            print(f"  {Colors.YELLOW}[1]{Colors.RESET} Excel (.xlsx)")
            print(f"  {Colors.YELLOW}[2]{Colors.RESET} TXT文本")
            print(f"  {Colors.YELLOW}[3]{Colors.RESET} 两种都要")
            print(f"  {Colors.YELLOW}[0]{Colors.RESET} 不导出")
            
            export_choice = input(f"\n{Colors.BOLD}选择: {Colors.RESET}").strip()
            
            if export_choice == '1':
                filename = self.crawler.save_excel(show_data, self.video_info['title'], keyword)
                if filename:
                    print_status(f"Excel: {filename}", 'success')
            elif export_choice == '2':
                filename = self.crawler.save_txt(show_data, self.video_info['title'], keyword)
                if filename:
                    print_status(f"TXT: {filename}", 'success')
            elif export_choice == '3':
                f1 = self.crawler.save_excel(show_data, self.video_info['title'], keyword)
                f2 = self.crawler.save_txt(show_data, self.video_info['title'], keyword)
                if f1:
                    print_status(f"Excel: {f1}", 'success')
                if f2:
                    print_status(f"TXT: {f2}", 'success')
            
            print_status("完成！", 'success')
        else:
            print_status("未获取到评论", 'error')
        
        input(f"\n{Colors.GRAY}按回车继续...{Colors.RESET}")
    
    def keyword_menu(self):
        clear_screen()
        print_banner()
        
        print(f"\n{Colors.CYAN}🔑 关键词筛选{Colors.RESET}")
        print(f"{Colors.GRAY}输入关键词，只导出包含该词的评论{Colors.RESET}")
        print(f"{Colors.GRAY}支持多个关键词，用 | 分隔，如: 特朗普|天坛{Colors.RESET}")
        
        keyword = input(f"\n{Colors.BOLD}关键词 (回车取消): {Colors.RESET}").strip()
        
        if keyword:
            self.keyword = keyword
            print_status(f"已设置关键词: {keyword}", 'success')
        else:
            if hasattr(self, 'keyword'):
                delattr(self, 'keyword')
            print_status("已取消关键词筛选", 'info')
        
        time.sleep(0.5)
    
    def show_login_status(self):
        clear_screen()
        print_banner()
        
        if self.user_info:
            print(f"\n{Colors.BOLD}登录信息:{Colors.RESET}")
            print(f"  用户: {Colors.PINK}{self.user_info['name']}{Colors.RESET}")
            print(f"  等级: Lv{self.user_info['level']}")
            print(f"  硬币: {self.user_info['coins']}")
            
            cookie_file = 'bilibili_cookie.txt'
            if os.path.exists(cookie_file):
                size = os.path.getsize(cookie_file)
                print(f"  Cookie: {Colors.GREEN}已配置{Colors.RESET} ({size}字节)")
            else:
                print(f"  Cookie: {Colors.RED}未找到{Colors.RESET}")
        else:
            print_status("未登录", 'warning')
            print(f"\n运行 {Colors.YELLOW}python cookie.py{Colors.RESET} 获取Cookie")
        
        input(f"\n{Colors.GRAY}按回车返回...{Colors.RESET}")
    
    def show_files(self):
        clear_screen()
        print_banner()
        
        files = [f for f in os.listdir('.') if f.startswith('B站评论_')]
        
        if files:
            print(f"\n{Colors.BOLD}📁 已保存的文件 ({len(files)}个):{Colors.RESET}\n")
            for i, f in enumerate(files, 1):
                size = os.path.getsize(f)
                size_str = f"{size/1024:.1f}KB" if size > 1024 else f"{size}B"
                icon = '📊' if f.endswith('.xlsx') else '📄'
                print(f"  {i}. {icon} {f} ({size_str})")
        else:
            print_status("暂无保存的文件", 'info')
        
        input(f"\n{Colors.GRAY}按回车返回...{Colors.RESET}")

if __name__ == '__main__':
    try:
        app = App()
        app.run()
    except KeyboardInterrupt:
        print(f"\n\n{Colors.PINK}👋 已退出{Colors.RESET}\n")
    except Exception as e:
        print(f"\n{Colors.RED}错误: {e}{Colors.RESET}")
        import traceback
        traceback.print_exc()
        input("\n按回车退出...")