#!/usr/bin/env python3
"""
B站评论爬虫 - GUI版（扫码登录 + 指数退避 + 智能延迟）
"""
import requests
import json
import re
import time
import os
import sys
import threading
from datetime import datetime
import random

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# GUI库
try:
    import tkinter as tk
    from tkinter import ttk, messagebox, scrolledtext, filedialog
    import tkinter.font as tkfont
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    print("警告: tkinter不可用，请安装python3-tk")


# ==================== 扫码登录模块 ====================
def generate_qrcode_ascii(url):
    """生成二维码的ASCII字符画"""
    try:
        import qrcode
        from PIL import Image
        import io
        
        qr = qrcode.QRCode(box_size=2, border=1)
        qr.add_data(url)
        qr.make(fit=True)
        
        matrix = qr.get_matrix()
        ascii_qr = []
        for row in matrix:
            line = ''.join(['██' if cell else '  ' for cell in row])
            ascii_qr.append(line)
        
        return '\n'.join(ascii_qr)
    except ImportError:
        return None


def get_qrcode_image(url, size=250):
    """获取二维码图片（用于GUI显示）"""
    try:
        import qrcode
        from PIL import Image, ImageTk
        
        qr = qrcode.QRCode(box_size=4, border=2)
        qr.add_data(url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        img = img.resize((size, size), Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(img)
    except ImportError:
        return None


def qrcode_login_gui(status_callback=None, qrcode_callback=None):
    """B站扫码登录，带回调的版本"""
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    })
    
    if status_callback:
        status_callback("正在获取二维码...")
    
    # 获取二维码
    resp = session.get('https://passport.bilibili.com/x/passport-login/web/qrcode/generate')
    data = resp.json()
    
    if data['code'] != 0:
        if status_callback:
            status_callback("获取二维码失败")
        return None
    
    qrcode_url = data['data']['url']
    qrcode_key = data['data']['qrcode_key']
    
    # 显示二维码
    if qrcode_callback:
        qrcode_callback(qrcode_url)
    
    if status_callback:
        status_callback("请使用B站APP扫描二维码")
    
    # 等待扫码
    while True:
        time.sleep(2)
        
        check_url = f'https://passport.bilibili.com/x/passport-login/web/qrcode/poll?qrcode_key={qrcode_key}'
        resp = session.get(check_url)
        result = resp.json()
        
        code = result['data'].get('code', -1)
        
        if code == 0:
            # 登录成功
            cookies = session.cookies.get_dict()
            cookie_str = '; '.join([f'{k}={v}' for k, v in cookies.items()])
            
            with open('bilibili_cookie.txt', 'w', encoding='utf-8') as f:
                f.write(cookie_str)
            
            # 验证登录信息
            resp = session.get('https://api.bilibili.com/x/web-interface/nav')
            user = resp.json()
            if user['code'] == 0:
                if status_callback:
                    status_callback(f"登录成功！欢迎: {user['data']['uname']}")
                return cookie_str
            else:
                if status_callback:
                    status_callback("登录成功，但获取用户信息失败")
                return cookie_str
                
        elif code == 86038:
            if status_callback:
                status_callback("二维码已过期，请重新获取")
            return None
        elif code == 86090:
            if status_callback:
                status_callback("已扫描，请在手机上确认登录...")
        # 86101 表示未扫描，继续等待


def try_auto_login(status_callback=None):
    """尝试自动登录"""
    cookie_file = 'bilibili_cookie.txt'
    
    if os.path.exists(cookie_file):
        if status_callback:
            status_callback("检测到已有Cookie，正在验证...")
        
        with open(cookie_file, 'r', encoding='utf-8') as f:
            cookie_str = f.read().strip()
        
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        session.headers['Cookie'] = cookie_str
        
        try:
            resp = session.get('https://api.bilibili.com/x/web-interface/nav', timeout=10)
            data = resp.json()
            if data['code'] == 0 and data['data']['isLogin']:
                if status_callback:
                    status_callback(f"Cookie有效！欢迎: {data['data']['uname']}")
                return cookie_str
            else:
                if status_callback:
                    status_callback("Cookie已过期，需要重新登录")
        except Exception as e:
            if status_callback:
                status_callback(f"Cookie验证失败: {e}")
    
    return None


# ==================== 爬虫核心类 ====================
class BilibiliComment:
    def __init__(self, cookie_file='bilibili_cookie.txt', log_callback=None):
        self.session = requests.Session()
        self.log_callback = log_callback
        self.cookie = self.load_cookie(cookie_file)
        
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0',
        ]
        
        self.update_headers()
        
        if self.cookie:
            self.session.headers['Cookie'] = self.cookie
        
        self.request_count = 0
        self.last_request_time = 0
        self.stop_flag = False
    
    def log(self, msg, level='info'):
        if self.log_callback:
            self.log_callback(msg, level)
    
    def update_headers(self):
        self.session.headers.update({
            'User-Agent': random.choice(self.user_agents),
            'Referer': 'https://www.bilibili.com',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache',
        })
    
    def exponential_backoff(self, attempt, base_delay=5, max_delay=300):
        delay = min(base_delay * (2 ** attempt), max_delay)
        jitter = random.uniform(0, delay * 0.3)
        return delay + jitter
    
    def smart_delay(self):
        base_delay = random.uniform(2, 5)
        
        if self.request_count > 0 and self.request_count % 50 == 0:
            extra_delay = random.uniform(3, 8)
            total_delay = base_delay + extra_delay
            self.log(f"已完成{self.request_count}次请求，休息{extra_delay:.1f}秒", 'info')
        else:
            total_delay = base_delay
        
        if self.last_request_time > 0:
            elapsed = time.time() - self.last_request_time
            if elapsed < 1:
                time.sleep(1 - elapsed)
        
        time.sleep(total_delay)
        self.last_request_time = time.time()
        self.request_count += 1
        
        return total_delay
    
    def load_cookie(self, filepath):
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                cookie = f.read().strip()
            return cookie
        return None
    
    def check_login(self):
        if not self.cookie:
            return None
        
        try:
            resp = self.session.get('https://api.bilibili.com/x/web-interface/nav', timeout=10)
            data = resp.json()
            if data['code'] == 0 and data['data']['isLogin']:
                return {
                    'name': data['data']['uname'],
                    'level': data['data']['level_info']['current_level'],
                    'coins': data['data']['money'],
                }
        except:
            pass
        return None
    
    def get_video_info(self, url):
        bv = re.search(r'BV[a-zA-Z0-9]+', url)
        bvid = bv.group() if bv else url
        
        for retry in range(3):
            try:
                self.update_headers()
                resp = self.session.get(f'https://api.bilibili.com/x/web-interface/view?bvid={bvid}', timeout=10)
                
                if resp.status_code == 412:
                    delay = self.exponential_backoff(retry)
                    self.log(f"请求被限制，等待 {delay:.1f} 秒后重试...", 'warning')
                    time.sleep(delay)
                    continue
                
                data = resp.json()
                
                if data['code'] == 0:
                    return {
                        'aid': data['data']['aid'],
                        'bvid': bvid,
                        'title': data['data']['title'],
                        'total': data['data']['stat']['reply'],
                        'like': data['data']['stat']['like'],
                        'view': data['data']['stat']['view'],
                        'danmaku': data['data']['stat']['danmaku'],
                    }
                else:
                    self.log(f"API错误: {data.get('message', '未知错误')}", 'error')
                    
            except Exception as e:
                self.log(f"获取视频信息失败 (重试{retry+1}/3): {e}", 'warning')
                time.sleep(self.exponential_backoff(retry))
        
        return None
    
    def fetch_comments_safe(self, url, max_retries=5):
        for attempt in range(max_retries):
            if self.stop_flag:
                return None
                
            try:
                self.update_headers()
                resp = self.session.get(url, timeout=15)
                
                if resp.status_code == 403 or resp.status_code == 412:
                    delay = self.exponential_backoff(attempt)
                    self.log(f"被限制访问 (状态码{resp.status_code})，等待 {delay:.1f} 秒后重试", 'warning')
                    time.sleep(delay)
                    continue
                
                if resp.status_code != 200:
                    if attempt < max_retries - 1:
                        delay = self.exponential_backoff(attempt)
                        self.log(f"HTTP {resp.status_code}，等待 {delay:.1f} 秒后重试", 'warning')
                        time.sleep(delay)
                        continue
                    return None
                
                if not resp.text or not resp.text.strip():
                    if attempt < max_retries - 1:
                        delay = self.exponential_backoff(attempt)
                        self.log(f"空响应，等待 {delay:.1f} 秒后重试", 'warning')
                        time.sleep(delay)
                        continue
                    return None
                
                data = resp.json()
                
                if isinstance(data, dict) and data.get('code') == -403:
                    self.log("被反爬虫拦截", 'error')
                    if attempt < max_retries - 1:
                        delay = self.exponential_backoff(attempt, base_delay=10)
                        self.log(f"等待 {delay:.1f} 秒后重试...", 'warning')
                        time.sleep(delay)
                        continue
                    return None
                
                return data
                
            except json.JSONDecodeError:
                self.log(f"JSON解析失败 (尝试 {attempt+1}/{max_retries})", 'warning')
                if attempt < max_retries - 1:
                    delay = self.exponential_backoff(attempt)
                    time.sleep(delay)
                else:
                    return None
                    
            except requests.Timeout:
                self.log(f"请求超时 (尝试 {attempt+1}/{max_retries})", 'warning')
                if attempt < max_retries - 1:
                    delay = self.exponential_backoff(attempt)
                    time.sleep(delay)
                    continue
                return None
                
            except requests.RequestException as e:
                self.log(f"网络错误 (尝试 {attempt+1}/{max_retries}): {e}", 'warning')
                if attempt < max_retries - 1:
                    delay = self.exponential_backoff(attempt)
                    time.sleep(delay)
                else:
                    return None
        
        return None
    
    def fetch_comments(self, aid, max_count=500, sort=1, keyword=None, 
                       progress_callback=None, page_callback=None):
        """爬取指定数量评论"""
        comments = []
        filtered = []
        page = 1
        failed_pages = 0
        max_failed = 5
        consecutive_empty = 0
        
        self.log(f"开始爬取评论，目标数量: {max_count}", 'info')
        
        while len(comments) < max_count and not self.stop_flag:
            params = {
                'type': 1,
                'oid': aid,
                'pn': page,
                'ps': 20,
                'sort': sort,
            }
            
            url = 'https://api.bilibili.com/x/v2/reply'
            for key, value in params.items():
                url += f'&{key}={value}' if '?' in url else f'?{key}={value}'
            
            result = self.fetch_comments_safe(url)
            
            if result is None:
                failed_pages += 1
                if failed_pages >= max_failed:
                    self.log("连续失败次数过多，停止爬取", 'error')
                    break
                page += 1
                continue
            
            failed_pages = 0
            
            if result.get('code') != 0:
                error_msg = result.get('message', '未知错误')
                self.log(f"API返回错误: {error_msg}", 'warning')
                break
            
            replies = result.get('data', {}).get('replies', [])
            if not replies:
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    self.log("已到达最后一页", 'info')
                    break
                page += 1
                continue
            else:
                consecutive_empty = 0
            
            for r in replies:
                if len(comments) >= max_count or self.stop_flag:
                    break
                    
                content = r.get('content', {}).get('message', '')
                if not content or not content.strip():
                    continue
                
                comment_data = {
                    '序号': len(comments) + 1,
                    '用户': r.get('member', {}).get('uname', '未知'),
                    '用户ID': r.get('member', {}).get('mid', 0),
                    '等级': r.get('member', {}).get('level_info', {}).get('current_level', 0),
                    '内容': content,
                    '点赞': r.get('like', 0),
                    '回复数': r.get('rcount', 0),
                    '时间': datetime.fromtimestamp(r.get('ctime', 0)).strftime('%Y-%m-%d %H:%M:%S'),
                }
                comments.append(comment_data)
                
                if keyword and keyword.lower() in content.lower():
                    filtered.append(comment_data)
            
            if progress_callback:
                progress_callback(len(comments), max_count)
            
            if page_callback:
                page_callback(page, len(comments))
            
            self.log(f"第{page}页完成，已获取 {len(comments)}/{max_count} 条", 'info')
            
            page += 1
            
            if not self.stop_flag and len(comments) < max_count:
                delay = self.smart_delay()
                self.log(f"等待 {delay:.1f} 秒后继续...", 'debug')
        
        return comments, filtered
    
    def fetch_all_comments(self, aid, sort=1, keyword=None, 
                           progress_callback=None, page_callback=None):
        """爬取所有评论"""
        comments = []
        filtered = []
        page = 1
        failed_pages = 0
        max_failed = 8
        consecutive_empty = 0
        total_comments = None
        
        self.log("开始爬取所有评论", 'info')
        
        while not self.stop_flag:
            params = {
                'type': 1,
                'oid': aid,
                'pn': page,
                'ps': 20,
                'sort': sort,
            }
            
            url = 'https://api.bilibili.com/x/v2/reply'
            for key, value in params.items():
                url += f'&{key}={value}' if '?' in url else f'?{key}={value}'
            
            result = self.fetch_comments_safe(url)
            
            if result is None:
                failed_pages += 1
                self.log(f"第{page}页请求失败 ({failed_pages}/{max_failed})", 'warning')
                if failed_pages >= max_failed:
                    self.log("连续失败次数过多，停止爬取", 'error')
                    break
                page += 1
                continue
            
            failed_pages = 0
            
            if result.get('code') != 0:
                error_msg = result.get('message', '未知错误')
                self.log(f"API返回错误: {error_msg}", 'warning')
                if result.get('code') in [-404, -400]:
                    self.log("已到达最后一页", 'info')
                    break
                page += 1
                continue
            
            # 获取总评论数
            if total_comments is None:
                total_comments = result.get('data', {}).get('page', {}).get('count', 0)
                if total_comments > 0:
                    self.log(f"总评论数: {total_comments:,}", 'info')
                else:
                    self.log("无法获取总评论数，继续爬取...", 'warning')
            
            replies = result.get('data', {}).get('replies', [])
            
            if not replies:
                consecutive_empty += 1
                if consecutive_empty >= 5:
                    self.log("已到达最后一页", 'info')
                    break
                self.log(f"第{page}页无评论数据", 'warning')
                page += 1
                continue
            else:
                consecutive_empty = 0
            
            new_comments = 0
            for r in replies:
                if self.stop_flag:
                    break
                    
                content = r.get('content', {}).get('message', '')
                if not content or not content.strip():
                    continue
                
                comment_data = {
                    '序号': len(comments) + 1,
                    '用户': r.get('member', {}).get('uname', '未知'),
                    '用户ID': r.get('member', {}).get('mid', 0),
                    '等级': r.get('member', {}).get('level_info', {}).get('current_level', 0),
                    '内容': content,
                    '点赞': r.get('like', 0),
                    '回复数': r.get('rcount', 0),
                    '时间': datetime.fromtimestamp(r.get('ctime', 0)).strftime('%Y-%m-%d %H:%M:%S'),
                }
                comments.append(comment_data)
                new_comments += 1
                
                if keyword and keyword.lower() in content.lower():
                    filtered.append(comment_data)
            
            self.log(f"第{page}页 +{new_comments}条 (总计{len(comments)}条)", 'info')
            
            if progress_callback:
                if total_comments:
                    progress_callback(len(comments), total_comments)
                else:
                    progress_callback(len(comments), len(comments) + 20)
            
            if page_callback:
                page_callback(page, len(comments))
            
            # 检查是否还有更多页
            total_pages = (result.get('data', {}).get('page', {}).get('count', 0) + 19) // 20
            if page >= total_pages and total_pages > 0:
                self.log(f"已获取全部 {total_pages} 页", 'success')
                break
            
            page += 1
            
            if not self.stop_flag:
                delay = self.smart_delay()
                self.log(f"等待 {delay:.1f} 秒后继续...", 'debug')
        
        return comments, filtered
    
    def stop(self):
        self.stop_flag = True
    
    def save_excel(self, comments, title, keyword=None, save_path=None):
        if not HAS_EXCEL:
            self.log("请安装openpyxl: pip install openpyxl", 'warning')
            return None
        
        if not comments:
            self.log("没有数据可导出", 'warning')
            return None
        
        safe_title = re.sub(r'[\\/:*?"<>|]', '_', title)[:40]
        prefix = f"筛选_{keyword}_" if keyword else ""
        filename = f"B站评论_{prefix}{safe_title}_{len(comments)}条.xlsx"
        
        if save_path:
            filename = os.path.join(save_path, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "评论数据"
        
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='FB7299', end_color='FB7299', fill_type='solid')
        
        headers = ['序号', '用户', '用户ID', '等级', '内容', '点赞数', '回复数', '时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row, c in enumerate(comments, 2):
            ws.cell(row=row, column=1, value=c['序号'])
            ws.cell(row=row, column=2, value=c['用户'])
            ws.cell(row=row, column=3, value=c['用户ID'])
            ws.cell(row=row, column=4, value=c['等级'])
            ws.cell(row=row, column=5, value=c['内容'])
            ws.cell(row=row, column=6, value=c['点赞'])
            ws.cell(row=row, column=7, value=c['回复数'])
            ws.cell(row=row, column=8, value=c['时间'])
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:H{len(comments)+1}'
        
        wb.save(filename)
        self.log(f"已保存到: {filename}", 'success')
        return filename
    
    def save_txt(self, comments, title, keyword=None, save_path=None):
        if not comments:
            return None
            
        safe_title = re.sub(r'[\\/:*?"<>|]', '_', title)[:40]
        prefix = f"筛选_{keyword}_" if keyword else ""
        filename = f"B站评论_{prefix}{safe_title}_{len(comments)}条.txt"
        
        if save_path:
            filename = os.path.join(save_path, filename)
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"视频: {title}\n")
            f.write(f"评论: {len(comments)}条\n")
            if keyword:
                f.write(f"关键词: {keyword}\n")
            f.write(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            
            for c in comments:
                f.write(f"{'─' * 60}\n")
                f.write(f"【{c['序号']}】{c['用户']} | Lv{c['等级']} | {c['时间']}\n")
                f.write(f"👍{c['点赞']}  💬{c['回复数']}\n")
                f.write(f"\n{c['内容']}\n\n")
        
        self.log(f"已保存到: {filename}", 'success')
        return filename


# ==================== GUI 应用 ====================
class BilibiliGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("B站评论爬虫 Pro v3.3")
        self.root.geometry("900x750")
        self.root.minsize(800, 600)
        
        # 设置图标和样式
        self.root.configure(bg='#f0f0f0')
        
        # 爬虫实例
        self.crawler = None
        self.video_info = None
        self.comments = []
        self.filtered = []
        self.is_crawling = False
        self.crawl_thread = None
        
        # 关键词
        self.keyword_var = tk.StringVar()
        
        # 登录状态
        self.logged_in = False
        self.user_info = None
        
        # 创建UI
        self.create_widgets()
        
        # 自动登录
        self.root.after(100, self.auto_login)
    
    def create_widgets(self):
        # 设置样式
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TLabel', background='#f0f0f0', font=('微软雅黑', 10))
        style.configure('TButton', font=('微软雅黑', 10))
        style.configure('TLabelframe', background='#f0f0f0', font=('微软雅黑', 10))
        style.configure('TLabelframe.Label', background='#f0f0f0', font=('微软雅黑', 10, 'bold'))
        
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # ===== 顶部：视频输入区域 =====
        video_frame = ttk.LabelFrame(main_frame, text="视频信息", padding="10")
        video_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 第一行：链接输入
        row1 = ttk.Frame(video_frame)
        row1.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(row1, text="视频链接/BV号:").pack(side=tk.LEFT, padx=(0, 10))
        self.url_entry = ttk.Entry(row1, font=('微软雅黑', 10))
        self.url_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.get_info_btn = ttk.Button(row1, text="获取信息", command=self.get_video_info)
        self.get_info_btn.pack(side=tk.RIGHT)
        
        # 第二行：视频信息显示
        self.video_info_frame = ttk.Frame(video_frame)
        self.video_info_frame.pack(fill=tk.X)
        
        self.video_title_var = tk.StringVar(value="未获取视频")
        title_label = ttk.Label(self.video_info_frame, textvariable=self.video_title_var, 
                                 font=('微软雅黑', 11, 'bold'), foreground='#FB7299')
        title_label.pack(anchor=tk.W, pady=(0, 5))
        
        stats_frame = ttk.Frame(self.video_info_frame)
        stats_frame.pack(fill=tk.X)
        
        self.view_var = tk.StringVar(value="播放: -")
        self.like_var = tk.StringVar(value="点赞: -")
        self.comment_count_var = tk.StringVar(value="评论: -")
        
        ttk.Label(stats_frame, textvariable=self.view_var).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(stats_frame, textvariable=self.like_var).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(stats_frame, textvariable=self.comment_count_var).pack(side=tk.LEFT)
        
        # ===== 中间：设置区域 =====
        settings_frame = ttk.LabelFrame(main_frame, text="爬取设置", padding="10")
        settings_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 第一行：排序和数量
        row1 = ttk.Frame(settings_frame)
        row1.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(row1, text="排序方式:").pack(side=tk.LEFT, padx=(0, 10))
        self.sort_var = tk.StringVar(value="hot")
        ttk.Radiobutton(row1, text="按热度", variable=self.sort_var, value="hot").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(row1, text="按时间", variable=self.sort_var, value="time").pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(row1, text="爬取数量:").pack(side=tk.LEFT, padx=(0, 10))
        self.count_var = tk.StringVar(value="all")
        ttk.Radiobutton(row1, text="全部", variable=self.count_var, value="all").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(row1, text="指定数量", variable=self.count_var, value="limit").pack(side=tk.LEFT)
        
        self.limit_entry = ttk.Entry(row1, width=8, state='disabled')
        self.limit_entry.pack(side=tk.LEFT, padx=(10, 0))
        
        # 绑定数量选择变化
        def on_count_change():
            if self.count_var.get() == "limit":
                self.limit_entry.config(state='normal')
            else:
                self.limit_entry.config(state='disabled')
        
        self.count_var.trace('w', lambda *_: on_count_change())
        
        # 第二行：关键词筛选
        row2 = ttk.Frame(settings_frame)
        row2.pack(fill=tk.X)
        
        ttk.Label(row2, text="关键词筛选:").pack(side=tk.LEFT, padx=(0, 10))
        self.keyword_entry = ttk.Entry(row2, textvariable=self.keyword_var, width=30)
        self.keyword_entry.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(row2, text="(多个关键词用 | 分隔)", foreground='gray').pack(side=tk.LEFT)
        
        # ===== 日志区域 =====
        log_frame = ttk.LabelFrame(main_frame, text="运行日志", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 日志文本框
        self.log_text = scrolledtext.ScrolledText(log_frame, height=12, font=('Consolas', 9),
                                                   wrap=tk.WORD, bg='#1e1e1e', fg='#d4d4d4')
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 配置日志颜色标签
        self.log_text.tag_config('info', foreground='#4ec9b0')
        self.log_text.tag_config('success', foreground='#6a9955')
        self.log_text.tag_config('error', foreground='#f48771')
        self.log_text.tag_config('warning', foreground='#dcdcaa')
        self.log_text.tag_config('debug', foreground='#858585')
        
        # ===== 底部：控制按钮和进度条 =====
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X)
        
        # 左侧：登录状态
        self.login_status_var = tk.StringVar(value="未登录")
        login_status_label = ttk.Label(bottom_frame, textvariable=self.login_status_var, 
                                        font=('微软雅黑', 9), foreground='gray')
        login_status_label.pack(side=tk.LEFT)
        
        # 右侧：按钮
        button_frame = ttk.Frame(bottom_frame)
        button_frame.pack(side=tk.RIGHT)
        
        self.login_btn = ttk.Button(button_frame, text="登录", command=self.show_login_window, width=8)
        self.login_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.start_btn = ttk.Button(button_frame, text="开始爬取", command=self.start_crawl, width=10)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.start_btn.config(state='disabled')
        
        self.stop_btn = ttk.Button(button_frame, text="停止", command=self.stop_crawl, width=8, state='disabled')
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.export_btn = ttk.Button(button_frame, text="导出结果", command=self.export_data, width=8, state='disabled')
        self.export_btn.pack(side=tk.LEFT)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                             maximum=100, mode='determinate')
        self.progress_bar.pack(fill=tk.X, pady=(10, 0))
        
        self.progress_label = ttk.Label(main_frame, text="", font=('微软雅黑', 9))
        self.progress_label.pack()
    
    def log(self, msg, level='info'):
        """添加日志"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_line = f"[{timestamp}] {msg}\n"
        
        def _log():
            self.log_text.insert(tk.END, log_line, level)
            self.log_text.see(tk.END)
            self.root.update_idletasks()
        
        self.root.after(0, _log)
    
    def update_progress(self, current, total):
        """更新进度条"""
        if total > 0:
            percent = (current / total) * 100
        else:
            percent = 0
        
        def _update():
            self.progress_var.set(percent)
            self.progress_label.config(text=f"{current}/{total} ({percent:.1f}%)")
            self.root.update_idletasks()
        
        self.root.after(0, _update)
    
    def auto_login(self):
        """自动登录"""
        self.log("正在检查登录状态...", 'info')
        cookie = try_auto_login(status_callback=self.log)
        
        if cookie:
            self.crawler = BilibiliComment(log_callback=self.log)
            user_info = self.crawler.check_login()
            if user_info:
                self.logged_in = True
                self.user_info = user_info
                self.login_status_var.set(f"已登录: {user_info['name']}")
                self.log(f"登录成功！欢迎: {user_info['name']} (Lv{user_info['level']})", 'success')
            else:
                self.login_status_var.set("未登录")
        else:
            self.login_status_var.set("未登录")
            self.log("未检测到有效Cookie，请点击「登录」按钮", 'warning')
        
        # 创建基础爬虫实例
        if not self.crawler:
            self.crawler = BilibiliComment(log_callback=self.log)
    
    def show_login_window(self):
        """显示登录窗口"""
        login_window = tk.Toplevel(self.root)
        login_window.title("B站扫码登录")
        login_window.geometry("350x500")
        login_window.resizable(False, False)
        login_window.configure(bg='white')
        
        # 居中
        login_window.transient(self.root)
        login_window.grab_set()
        
        # 标题
        title_label = tk.Label(login_window, text="请使用B站APP扫描二维码", 
                                font=('微软雅黑', 14), bg='white', fg='#FB7299')
        title_label.pack(pady=(20, 10))
        
        # 二维码显示区域
        qr_frame = tk.Frame(login_window, bg='white', width=280, height=280)
        qr_frame.pack(pady=10)
        qr_frame.pack_propagate(False)
        
        qr_label = tk.Label(qr_frame, bg='white')
        qr_label.pack(fill=tk.BOTH, expand=True)
        
        # 状态标签
        status_var = tk.StringVar(value="正在获取二维码...")
        status_label = tk.Label(login_window, textvariable=status_var, 
                                 font=('微软雅黑', 10), bg='white', fg='gray')
        status_label.pack(pady=10)
        
        # 取消按钮
        cancel_btn = ttk.Button(login_window, text="取消", 
                                command=login_window.destroy)
        cancel_btn.pack(pady=10)
        
        def on_status(msg):
            status_var.set(msg)
        
        def on_qrcode(url):
            img = get_qrcode_image(url, size=250)
            if img:
                qr_label.config(image=img)
                qr_label.image = img
                status_var.set("请扫描二维码")
            else:
                status_var.set("请安装 qrcode 和 pillow 库")
        
        def do_login():
            cookie = qrcode_login_gui(status_callback=on_status, qrcode_callback=on_qrcode)
            if cookie:
                status_var.set("登录成功！")
                self.root.after(500, login_window.destroy)
                self.root.after(500, lambda: self.reload_crawler())
            else:
                status_var.set("登录失败，请重试")
        
        # 在新线程中执行登录
        thread = threading.Thread(target=do_login, daemon=True)
        thread.start()
    
    def reload_crawler(self):
        """重新加载爬虫实例"""
        self.crawler = BilibiliComment(log_callback=self.log)
        user_info = self.crawler.check_login()
        if user_info:
            self.logged_in = True
            self.user_info = user_info
            self.login_status_var.set(f"已登录: {user_info['name']}")
            self.log(f"登录成功！欢迎: {user_info['name']} (Lv{user_info['level']})", 'success')
        else:
            self.log("登录验证失败", 'error')
    
    def get_video_info(self):
        """获取视频信息"""
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showwarning("提示", "请输入视频链接或BV号")
            return
        
        self.log(f"正在获取视频信息: {url}", 'info')
        self.get_info_btn.config(text="获取中...", state='disabled')
        
        def fetch():
            crawler = BilibiliComment(log_callback=self.log)
            video_info = crawler.get_video_info(url)
            
            def update():
                self.get_info_btn.config(text="获取信息", state='normal')
                if video_info:
                    self.video_info = video_info
                    self.video_title_var.set(video_info['title'])
                    self.view_var.set(f"播放: {video_info['view']:,}")
                    self.like_var.set(f"点赞: {video_info['like']:,}")
                    self.comment_count_var.set(f"评论: {video_info['total']:,}")
                    self.log(f"获取成功: {video_info['title']}", 'success')
                    self.start_btn.config(state='normal')
                else:
                    self.log("获取视频信息失败", 'error')
                    messagebox.showerror("错误", "获取视频信息失败，请检查链接")
            
            self.root.after(0, update)
        
        thread = threading.Thread(target=fetch, daemon=True)
        thread.start()
    
    def start_crawl(self):
        """开始爬取"""
        if not self.video_info:
            messagebox.showwarning("提示", "请先获取视频信息")
            return
        
        if self.is_crawling:
            messagebox.showwarning("提示", "爬取正在进行中")
            return
        
        # 获取设置
        sort = 2 if self.sort_var.get() == "time" else 1
        
        max_count = None
        if self.count_var.get() == "limit":
            try:
                max_count = int(self.limit_entry.get())
                if max_count <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("提示", "请输入有效的数量")
                return
        
        keyword = self.keyword_var.get().strip()
        if not keyword:
            keyword = None
        
        # 重置状态
        self.comments = []
        self.filtered = []
        self.is_crawling = True
        self.start_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.export_btn.config(state='disabled')
        self.progress_var.set(0)
        self.progress_label.config(text="准备开始...")
        
        self.log(f"开始爬取视频: {self.video_info['title']}", 'info')
        if keyword:
            self.log(f"关键词筛选: {keyword}", 'info')
        
        def crawl():
            try:
                if max_count:
                    comments, filtered = self.crawler.fetch_comments(
                        self.video_info['aid'],
                        max_count,
                        sort,
                        keyword,
                        progress_callback=self.update_progress,
                        page_callback=None
                    )
                else:
                    comments, filtered = self.crawler.fetch_all_comments(
                        self.video_info['aid'],
                        sort,
                        keyword,
                        progress_callback=self.update_progress,
                        page_callback=None
                    )
                
                self.comments = comments
                self.filtered = filtered
                
                def finish():
                    self.is_crawling = False
                    self.start_btn.config(state='normal')
                    self.stop_btn.config(state='disabled')
                    
                    if comments:
                        self.export_btn.config(state='normal')
                        self.log(f"爬取完成！共 {len(comments):,} 条评论", 'success')
                        if keyword:
                            self.log(f"匹配关键词: {len(filtered):,} 条", 'info')
                        
                        # 统计
                        show_data = filtered if keyword else comments
                        if show_data:
                            likes = [c['点赞'] for c in show_data]
                            self.log(f"最多点赞: {max(likes):,}", 'info')
                            self.log(f"平均点赞: {sum(likes)//len(likes):,}", 'info')
                    else:
                        self.log("未获取到任何评论", 'warning')
                    
                    self.progress_label.config(text="完成")
                    self.update_progress(100, 100)
                
                self.root.after(0, finish)
                
            except Exception as e:
                self.log(f"爬取出错: {e}", 'error')
                
                def error_finish():
                    self.is_crawling = False
                    self.start_btn.config(state='normal')
                    self.stop_btn.config(state='disabled')
                    self.progress_label.config(text="出错")
                
                self.root.after(0, error_finish)
        
        thread = threading.Thread(target=crawl, daemon=True)
        thread.start()
    
    def stop_crawl(self):
        """停止爬取"""
        if self.crawler:
            self.crawler.stop()
            self.log("正在停止爬取...", 'warning')
        
        self.is_crawling = False
        self.start_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
    
    def export_data(self):
        """导出数据"""
        show_data = self.filtered if self.keyword_var.get() else self.comments
        
        if not show_data:
            messagebox.showwarning("提示", "没有数据可导出")
            return
        
        # 选择保存目录
        save_dir = filedialog.askdirectory(title="选择保存目录")
        if not save_dir:
            return
        
        # 导出格式选择
        result = messagebox.askquestion("导出格式", "是否同时导出Excel和TXT？\n\n选择「是」导出两种格式\n选择「否」仅导出Excel", 
                                        icon='question')
        
        self.log("正在导出数据...", 'info')
        
        def export():
            try:
                if result == 'yes':
                    f1 = self.crawler.save_excel(show_data, self.video_info['title'], 
                                                  self.keyword_var.get(), save_dir)
                    f2 = self.crawler.save_txt(show_data, self.video_info['title'],
                                                self.keyword_var.get(), save_dir)
                    if f1:
                        self.log(f"Excel已保存: {f1}", 'success')
                    if f2:
                        self.log(f"TXT已保存: {f2}", 'success')
                else:
                    f1 = self.crawler.save_excel(show_data, self.video_info['title'],
                                                  self.keyword_var.get(), save_dir)
                    if f1:
                        self.log(f"Excel已保存: {f1}", 'success')
                
                self.root.after(0, lambda: messagebox.showinfo("完成", "导出完成！"))
            except Exception as e:
                self.log(f"导出失败: {e}", 'error')
                self.root.after(0, lambda: messagebox.showerror("错误", f"导出失败: {e}"))
        
        thread = threading.Thread(target=export, daemon=True)
        thread.start()
    
    def run(self):
        self.root.mainloop()


# ==================== 入口 ====================
def main():
    if not GUI_AVAILABLE:
        print("错误: tkinter不可用，请安装python3-tk")
        print("Ubuntu/Debian: sudo apt install python3-tk")
        print("Windows: 重新安装Python并确保勾选tcl/tk")
        return
    
    app = BilibiliGUI()
    app.run()


if __name__ == '__main__':
    main()